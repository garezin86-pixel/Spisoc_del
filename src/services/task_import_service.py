# src/services/task_import_service.py
import csv
import io
from dataclasses import dataclass
from datetime import datetime

from openpyxl import load_workbook

from src.core.task_labels import PRIORITY_LABELS
from src.models.enums import TaskPriority

# Заголовки, по которым ищем нужные колонки — регистронезависимо, с алиасами,
# чтобы можно было импортировать и файл, экспортированный этой же системой
# (там есть ID/Статус/Автор и т.д. — их просто игнорируем), и файл из другого
# трекера с произвольным набором колонок.
TITLE_ALIASES = {"название", "title", "name", "задача", "task"}
DEADLINE_ALIASES = {"дедлайн", "deadline", "срок", "due date", "due"}
PRIORITY_ALIASES = {"приоритет", "priority"}

# PRIORITY_LABELS маппит строковое значение enum'а ("medium", "high", ...) на
# русскую подпись ("Средний", "Высокий", ...) — см. использование в
# task_export_service.py: PRIORITY_LABELS.get(task.priority.value, ...).
# Поэтому здесь конвертируем строковый ключ обратно в TaskPriority через
# TaskPriority(value), а не берём его как есть.
_LABEL_TO_PRIORITY: dict[str, TaskPriority] = {
    label.strip().lower(): TaskPriority(value) for value, label in PRIORITY_LABELS.items()
}
_VALUE_TO_PRIORITY: dict[str, TaskPriority] = {p.value.lower(): p for p in TaskPriority}

_DEADLINE_FORMATS = (
    "%d.%m.%Y %H:%M",
    "%d.%m.%Y",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
)

MAX_IMPORT_ROWS = 2000


class TaskImportParseError(Exception):
    """Файл в целом не удалось разобрать (не тот формат, нет нужных колонок и т.п.)."""


@dataclass
class ParsedTaskRow:
    row_number: int  # номер строки в исходном файле (для обратной связи пользователю)
    title: str
    deadline: datetime | None
    priority: TaskPriority


@dataclass
class RowIssue:
    row_number: int
    message: str


@dataclass
class ImportParseResult:
    rows: list[ParsedTaskRow]
    errors: list[RowIssue]  # строка полностью пропущена (например, пустое название)
    warnings: list[RowIssue]  # строка создана, но с оговоркой (например, приоритет не распознан)


def _normalize_header(value) -> str:
    return str(value or "").strip().lower()


def _resolve_columns(header_row: list) -> dict[str, int]:
    """Находит индексы колонок Название/Дедлайн/Приоритет по алиасам."""
    columns: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        header = _normalize_header(raw)
        if header in TITLE_ALIASES and "title" not in columns:
            columns["title"] = idx
        elif header in DEADLINE_ALIASES and "deadline" not in columns:
            columns["deadline"] = idx
        elif header in PRIORITY_ALIASES and "priority" not in columns:
            columns["priority"] = idx

    if "title" not in columns:
        raise TaskImportParseError("Не найдена колонка с названием задачи (ожидается «Название» или «Title»)")
    return columns


def _parse_deadline(raw) -> tuple[datetime | None, str | None]:
    """Возвращает (deadline, error). error не None, если значение было, но не распозналось."""
    if raw is None or raw == "":
        return None, None
    if isinstance(raw, datetime):
        return raw, None

    text = str(raw).strip()
    for fmt in _DEADLINE_FORMATS:
        try:
            return datetime.strptime(text, fmt), None
        except ValueError:
            continue
    return None, f"дедлайн «{text}» не распознан, оставлен пустым"


def _parse_priority(raw) -> tuple[TaskPriority, str | None]:
    """Возвращает (priority, warning). При нераспознанном значении — medium + предупреждение."""
    if raw is None or str(raw).strip() == "":
        return TaskPriority.medium, None

    key = str(raw).strip().lower()
    priority = _LABEL_TO_PRIORITY.get(key) or _VALUE_TO_PRIORITY.get(key)
    if priority is None:
        return (
            TaskPriority.medium,
            f"приоритет «{raw}» не распознан, поставлен «Средний»",
        )
    return priority, None


def _rows_from_csv(content: bytes) -> tuple[list, list[list]]:
    # Excel на Windows часто сохраняет CSV в cp1251 без BOM; свой же экспорт —
    # utf-8 с BOM. Пробуем варианты по очереди.
    text = None
    for encoding in ("utf-8-sig", "cp1251", "utf-8"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise TaskImportParseError("Не удалось определить кодировку файла")

    # Свой экспорт — через запятую; файлы из Excel (сохранение "как есть")
    # часто через точку с запятой при локали ru_RU. Определяем по первой строке.
    lines = text.splitlines()
    first_line = lines[0] if lines else ""
    delimiter = ";" if first_line.count(";") > first_line.count(",") else ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        raise TaskImportParseError("Файл пуст")
    return rows[0], rows[1:]


def _rows_from_xlsx(content: bytes) -> tuple[list, list[list]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl кидает разные исключения на битых файлах
        raise TaskImportParseError("Не удалось открыть Excel-файл — возможно, он повреждён") from exc

    sheet = workbook.active
    if sheet is None:
        raise TaskImportParseError("В книге нет ни одного листа")
    all_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    if not all_rows:
        raise TaskImportParseError("Файл пуст")
    return all_rows[0], all_rows[1:]


def parse_import_file(filename: str, content: bytes) -> ImportParseResult:
    """Разбирает CSV/XLSX в список задач, готовых к созданию.

    Построчные проблемы не прерывают импорт целиком:
      - пустое название -> строка пропускается (errors)
      - нераспознанный дедлайн -> дедлайн пустой, задача создаётся (warnings)
      - нераспознанный приоритет -> priority=medium, задача создаётся (warnings)

    Прерывает импорт целиком (TaskImportParseError): неверный формат файла,
    отсутствие колонки с названием, повреждённый файл, файл пуст, либо строк
    больше MAX_IMPORT_ROWS.
    """
    name = filename.lower()
    if name.endswith(".xlsx"):
        header, data_rows = _rows_from_xlsx(content)
    elif name.endswith(".csv"):
        header, data_rows = _rows_from_csv(content)
    else:
        raise TaskImportParseError("Поддерживаются только файлы .csv и .xlsx")

    if len(data_rows) > MAX_IMPORT_ROWS:
        raise TaskImportParseError(f"Слишком много строк ({len(data_rows)}), максимум {MAX_IMPORT_ROWS} за один импорт")

    columns = _resolve_columns(header)

    rows: list[ParsedTaskRow] = []
    errors: list[RowIssue] = []
    warnings: list[RowIssue] = []

    for offset, raw_row in enumerate(data_rows):
        row_number = offset + 2  # +1 за 0-индексацию, +1 за строку заголовка
        title = str(raw_row[columns["title"]] or "").strip() if columns["title"] < len(raw_row) else ""

        if not title:
            errors.append(RowIssue(row_number, "пустое название — строка пропущена"))
            continue

        deadline_raw = (
            raw_row[columns["deadline"]] if "deadline" in columns and columns["deadline"] < len(raw_row) else None
        )
        priority_raw = (
            raw_row[columns["priority"]] if "priority" in columns and columns["priority"] < len(raw_row) else None
        )

        deadline, deadline_warning = _parse_deadline(deadline_raw)
        priority, priority_warning = _parse_priority(priority_raw)

        if deadline_warning:
            warnings.append(RowIssue(row_number, deadline_warning))
        if priority_warning:
            warnings.append(RowIssue(row_number, priority_warning))

        rows.append(ParsedTaskRow(row_number=row_number, title=title, deadline=deadline, priority=priority))

    return ImportParseResult(rows=rows, errors=errors, warnings=warnings)
