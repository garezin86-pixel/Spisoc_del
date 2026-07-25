# tests/test_task_import_parser.py
"""
Unit-тесты src.services.task_import_service — чистый парсинг, без БД.
Зеркально к тестам экспорта (test_task_export.py), только в обратную сторону.
"""

import io

import pytest
from openpyxl import Workbook

from src.models.enums import TaskPriority
from src.services.task_import_service import (
    MAX_IMPORT_ROWS,
    TaskImportParseError,
    parse_import_file,
)


def csv_bytes(text: str, encoding: str = "utf-8-sig") -> bytes:
    return text.encode(encoding)


def xlsx_bytes(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestFileFormatDetection:
    def test_unsupported_extension_raises(self):
        with pytest.raises(TaskImportParseError, match="csv.*xlsx|xlsx.*csv"):
            parse_import_file("tasks.txt", "Название\nX".encode("utf-8"))

    def test_empty_csv_raises(self):
        with pytest.raises(TaskImportParseError, match="пуст"):
            parse_import_file("tasks.csv", b"")

    def test_no_active_sheet_raises(self, monkeypatch):
        """openpyxl не даёт сохранить книгу без видимых листов (сам падает
        IndexError'ом при wb.save()), поэтому такой файл нельзя получить
        честным сохранением — подменяем load_workbook, чтобы напрямую
        проверить ветку `sheet is None` в _rows_from_xlsx."""

        class FakeWorkbook:
            active = None

        monkeypatch.setattr(
            "src.services.task_import_service.load_workbook",
            lambda *args, **kwargs: FakeWorkbook(),
        )
        with pytest.raises(TaskImportParseError):
            parse_import_file("tasks.xlsx", b"fake bytes, load_workbook is mocked")

    def test_corrupted_xlsx_raises(self):
        with pytest.raises(TaskImportParseError, match="повреждён"):
            parse_import_file("tasks.xlsx", b"this is not a real xlsx file")


class TestColumnResolution:
    def test_missing_title_column_raises(self):
        content = csv_bytes("Дедлайн,Приоритет\n01.01.2030,high\n")
        with pytest.raises(TaskImportParseError, match="[Нн]азвание"):
            parse_import_file("tasks.csv", content)

    def test_title_column_alias_english(self):
        content = csv_bytes("Title,Deadline,Priority\nTask A,,\n")
        result = parse_import_file("tasks.csv", content)
        assert len(result.rows) == 1
        assert result.rows[0].title == "Task A"

    def test_extra_columns_ignored(self):
        """Файл, экспортированный этой же системой (ID/Статус/Автор и т.д.),
        должен успешно импортироваться — лишние колонки просто игнорируются."""
        content = csv_bytes(
            "ID,Название,Описание,Статус,Приоритет,Автор,Исполнитель,Группа,Проект,Дедлайн,Создано,Теги,Повторение\n"
            "1,Экспортированная задача,desc,В работе,high,admin,admin,,,,,,\n"
        )
        result = parse_import_file("tasks.csv", content)
        assert len(result.rows) == 1
        assert result.rows[0].title == "Экспортированная задача"

    def test_column_order_does_not_matter(self):
        content = csv_bytes("Приоритет,Название,Дедлайн\nhigh,Задача,\n")
        result = parse_import_file("tasks.csv", content)
        assert result.rows[0].title == "Задача"
        assert result.rows[0].priority == TaskPriority.high

    def test_column_names_case_insensitive(self):
        content = csv_bytes("НАЗВАНИЕ\nЗадача в верхнем регистре\n")
        result = parse_import_file("tasks.csv", content)
        assert result.rows[0].title == "Задача в верхнем регистре"


class TestEncodingAndDelimiter:
    def test_utf8_bom_comma_delimited(self):
        content = csv_bytes("Название,Приоритет\nЗадача,high\n", encoding="utf-8-sig")
        result = parse_import_file("tasks.csv", content)
        assert result.rows[0].title == "Задача"

    def test_cp1251_semicolon_delimited(self):
        # Типичный экспорт из Excel на Windows с локалью ru_RU
        text = "Название;Приоритет\nЗадача из Excel;medium\n"
        content = text.encode("cp1251")
        result = parse_import_file("tasks.csv", content)
        assert result.rows[0].title == "Задача из Excel"
        assert result.rows[0].priority == TaskPriority.medium

    def test_plain_utf8_no_bom(self):
        content = "Название\nЗадача без BOM\n".encode("utf-8")
        result = parse_import_file("tasks.csv", content)
        assert result.rows[0].title == "Задача без BOM"


class TestXlsxParsing:
    def test_basic_xlsx_import(self):
        content = xlsx_bytes(
            [
                ["Название", "Дедлайн", "Приоритет"],
                ["Задача из Excel", "01.01.2030", "high"],
            ]
        )
        result = parse_import_file("tasks.xlsx", content)
        assert len(result.rows) == 1
        assert result.rows[0].title == "Задача из Excel"
        assert result.rows[0].priority == TaskPriority.high

    def test_xlsx_native_datetime_deadline(self):
        """openpyxl отдаёт datetime-объект для дат, а не строку — парсер должен
        принимать его напрямую, без strptime."""
        from datetime import datetime

        content = xlsx_bytes(
            [
                ["Название", "Дедлайн"],
                ["Задача", datetime(2030, 6, 15, 10, 0)],
            ]
        )
        result = parse_import_file("tasks.xlsx", content)
        assert result.rows[0].deadline == datetime(2030, 6, 15, 10, 0)


class TestRowLevelValidation:
    def test_empty_title_row_skipped_as_error(self):
        content = csv_bytes("Название\nЗадача 1\n\nЗадача 2\n")
        result = parse_import_file("tasks.csv", content)
        titles = [r.title for r in result.rows]
        assert titles == ["Задача 1", "Задача 2"]
        assert len(result.errors) == 1
        assert result.errors[0].row_number == 3  # заголовок=1, "Задача 1"=2, пустая=3

    def test_unrecognized_deadline_becomes_warning_not_error(self):
        content = csv_bytes("Название,Дедлайн\nЗадача,не дата вообще\n")
        result = parse_import_file("tasks.csv", content)
        assert len(result.rows) == 1
        assert result.rows[0].deadline is None
        assert len(result.warnings) == 1
        assert len(result.errors) == 0

    def test_unrecognized_priority_defaults_to_medium_with_warning(self):
        content = csv_bytes("Название,Приоритет\nЗадача,очень срочно\n")
        result = parse_import_file("tasks.csv", content)
        assert result.rows[0].priority == TaskPriority.medium
        assert len(result.warnings) == 1

    def test_empty_priority_defaults_to_medium_silently(self):
        """Пустое значение — не ошибка пользователя, просто дефолт без warning."""
        content = csv_bytes("Название,Приоритет\nЗадача,\n")
        result = parse_import_file("tasks.csv", content)
        assert result.rows[0].priority == TaskPriority.medium
        assert len(result.warnings) == 0

    def test_priority_recognized_by_russian_label(self):
        from src.core.task_labels import PRIORITY_LABELS

        label = PRIORITY_LABELS["critical"]
        content = f"Название,Приоритет\nЗадача,{label}\n".encode("utf-8-sig")
        result = parse_import_file("tasks.csv", content)
        assert result.rows[0].priority == TaskPriority.critical

    def test_priority_recognized_by_english_value(self):
        content = csv_bytes("Название,Приоритет\nЗадача,critical\n")
        result = parse_import_file("tasks.csv", content)
        assert result.rows[0].priority == TaskPriority.critical

    def test_deadline_formats_recognized(self):
        content = csv_bytes("Название,Дедлайн\nA,31.12.2030 23:59\nB,31.12.2030\nC,2030-12-31 23:59:00\nD,2030-12-31\n")
        result = parse_import_file("tasks.csv", content)
        assert len(result.rows) == 4
        assert all(r.deadline is not None for r in result.rows)

    def test_empty_deadline_is_none_without_warning(self):
        content = csv_bytes("Название,Дедлайн\nЗадача,\n")
        result = parse_import_file("tasks.csv", content)
        assert result.rows[0].deadline is None
        assert len(result.warnings) == 0

    def test_no_deadline_column_at_all(self):
        """Дедлайн и приоритет — необязательные колонки, файл только с
        Название должен успешно импортироваться."""
        content = csv_bytes("Название\nЗадача без остальных колонок\n")
        result = parse_import_file("tasks.csv", content)
        assert len(result.rows) == 1
        assert result.rows[0].deadline is None
        assert result.rows[0].priority == TaskPriority.medium

    def test_whitespace_only_title_treated_as_empty(self):
        content = csv_bytes('Название\n"   "\n')
        result = parse_import_file("tasks.csv", content)
        assert len(result.rows) == 0
        assert len(result.errors) == 1


class TestRowLimit:
    def test_too_many_rows_raises(self):
        header = "Название\n"
        body = "\n".join(f"Задача {i}" for i in range(MAX_IMPORT_ROWS + 1))
        content = (header + body).encode("utf-8-sig")
        with pytest.raises(TaskImportParseError, match="[Сс]лишком много"):
            parse_import_file("tasks.csv", content)

    def test_exactly_max_rows_ok(self):
        header = "Название\n"
        body = "\n".join(f"Задача {i}" for i in range(MAX_IMPORT_ROWS))
        content = (header + body).encode("utf-8-sig")
        result = parse_import_file("tasks.csv", content)
        assert len(result.rows) == MAX_IMPORT_ROWS
